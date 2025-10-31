# calculadora_tela_v2.py
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import math
from datetime import datetime

# openpyxl es opcional (se usa si el usuario quiere guardar en .xlsx)
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None


# ------------------------
# UTILIDADES NUMÉRICAS
# ------------------------
def format_number(n):
    """Devuelve el número como entero si no tiene decimales significativos."""
    if isinstance(n, float) and n.is_integer():
        return int(n)
    return n

def cm_to_m_str(cm):
    """Devuelve string 'X cm (Y m)' con 2 decimales en metros."""
    try:
        m = float(cm) / 100.0
        return f"{format_number(float(cm))} cm ({m:.2f} m)"
    except Exception:
        return ""

def format_cost(n):
    """Formatea un número para la visualización de costos."""
    if isinstance(n, str):
        return n
    if isinstance(n, float) and n.is_integer():
        return int(n)
    return f"{n:.2f}"


# ------------------------
# LÓGICA DE CÁLCULOS
# ------------------------
def calcular_tela_por_cantidad(ancho_tela_cm, ancho_molde_cm, alto_molde_cm,
                               margen_costura_cm, desperdicio_pct, cantidad, doble_molde=False):
    """
    Calcula la tela necesaria (largo en cm) para producir 'cantidad' piezas.
    Si doble_molde=True entonces la cantidad se multiplica por 2 (frente+contrafrente).
    Aplica margen por lado (se suma 2*margen al ancho y al alto).
    Aplica desperdicio (%) sobre el largo total final.
    Devuelve dict con resultados o (None, error_msg).
    """
    if doble_molde:
        cantidad = cantidad * 2

    ancho_total = ancho_molde_cm + 2 * margen_costura_cm
    alto_total = alto_molde_cm + 2 * margen_costura_cm

    if ancho_total <= 0 or alto_total <= 0:
        return None, "Dimensiones o margen inválidos."
    if ancho_total > ancho_tela_cm:
        return None, "El ancho total del molde (incluido margen) supera el ancho utilizable de la tela."

    moldes_por_fila = int(math.floor(ancho_tela_cm / ancho_total))
    if moldes_por_fila <= 0:
        return None, "No entra ningún molde por fila."

    filas_necesarias = int(math.ceil(cantidad / moldes_por_fila))
    largo_total_sin_desperdicio = filas_necesarias * alto_total
    largo_total_con_desperdicio = largo_total_sin_desperdicio * (1 + desperdicio_pct / 100.0)

    res = {
        "modo": "Calcular tela según cantidad de objetos/piezas",
        "ancho_tela_cm": ancho_tela_cm,
        "ancho_molde_cm": ancho_molde_cm,
        "alto_molde_cm": alto_molde_cm,
        "margen_costura_cm_por_lado": margen_costura_cm,
        "ancho_molde_total_cm": ancho_total,
        "alto_molde_total_cm": alto_total,
        "moldes_por_fila": moldes_por_fila,
        "filas_necesarias": filas_necesarias,
        "cantidad_solicitada": cantidad,
        "doble_molde": doble_molde,
        "largo_total_sin_desperdicio_cm": round(largo_total_sin_desperdicio, 2),
        "largo_total_con_desperdicio_cm": round(largo_total_con_desperdicio, 2)
    }
    return res, None


def calcular_moldes_con_tela(ancho_tela_cm, ancho_molde_cm, alto_molde_cm,
                             margen_costura_cm, desperdicio_pct, largo_tela_disponible_cm):
    """
    Calcula cuántos moldes se obtienen con un largo de tela disponible (en cm).
    El desperdicio (%) reduce la longitud utilizable: se considera que el largo real utilizable
    es largo_tela_disponible_cm / (1 + desperdicio_pct/100).
    """
    ancho_total = ancho_molde_cm + 2 * margen_costura_cm
    alto_total = alto_molde_cm + 2 * margen_costura_cm

    if ancho_total <= 0 or alto_total <= 0:
        return None, "Dimensiones o margen inválidos."
    if ancho_total > ancho_tela_cm:
        return None, "El ancho total del molde (incluido margen) supera el ancho utilizable de la tela."

    moldes_por_fila = int(math.floor(ancho_tela_cm / ancho_total))
    if moldes_por_fila <= 0:
        return None, "No entra ningún molde por fila."

    # descontar desperdicio: largo utilizable real
    if (1 + desperdicio_pct / 100.0) <= 0:
        return None, "Porcentaje de desperdicio inválido."
    largo_utilizable_cm = largo_tela_disponible_cm / (1 + desperdicio_pct / 100.0)

    filas_posibles = int(math.floor(largo_utilizable_cm / alto_total))
    total_moldes = moldes_por_fila * filas_posibles

    res = {
        "modo": "Calcular moldes con X cm de tela",
        "ancho_tela_cm": ancho_tela_cm,
        "ancho_molde_cm": ancho_molde_cm,
        "alto_molde_cm": alto_molde_cm,
        "margen_costura_cm_por_lado": margen_costura_cm,
        "ancho_molde_total_cm": ancho_total,
        "alto_molde_total_cm": alto_total,
        "moldes_por_fila": moldes_por_fila,
        "filas_posibles": filas_posibles,
        "total_moldes_obtenibles": total_moldes,
        "largo_tela_disponible_cm": largo_tela_disponible_cm,
        "largo_utilizable_cm": round(largo_utilizable_cm, 2)
    }
    return res, None


def calcular_costos_desde_largo(largo_tela_cm, precio_por_metro, cantidad_unidades=None):
    """
    Devuelve dict con precio_por_cm, costo_total y costo_unitario (si cantidad_unidades se pasa).
    """
    precio_por_cm = precio_por_metro / 100.0
    costo_total = precio_por_cm * largo_tela_cm
    costo_unitario = None
    if cantidad_unidades and cantidad_unidades > 0:
        costo_unitario = costo_total / cantidad_unidades
    return {
        "precio_por_cm": round(precio_por_cm, 6),
        "costo_total": round(costo_total, 4),
        "costo_unitario": (round(costo_unitario, 6) if costo_unitario is not None else "N/A")
    }


# ------------------------
# GUARDADO (TXT / XLSX)
# ------------------------
def guardar_txt(filepath, resumen):
    with open(filepath, "w", encoding="utf-8") as f:
        f.write("RESULTADO - " + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n\n")
        for campo, valor in resumen.items():
            f.write(f"{campo}: {valor}\n")


def guardar_xlsx(filepath, resumen):
    if openpyxl is None:
        raise RuntimeError("openpyxl no está instalado. Instala con: pip install openpyxl")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws.cell(row=1, column=1, value="Campo")
    ws.cell(row=1, column=2, value="Valor")
    r = 2
    for campo, valor in resumen.items():
        ws.cell(row=r, column=1, value=str(campo))
        ws.cell(row=r, column=2, value=str(valor))
        r += 1
    # ajustar anchos
    for col in [1, 2]:
        max_len = 0
        for cell in ws[get_column_letter(col)]:
            v = cell.value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = l
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)
    wb.save(filepath)


# ------------------------
# APLICACIÓN / UI
# ------------------------
class CalculadoraTelaApp:
    def __init__(self, master):
        self.master = master
        master.title("Calculadora de Tela – Producción Textil")
        master.geometry("880x660")
        master.resizable(False, False)

        # color beige amarillito para recuadros principales
        self.panel_bg = "#f9f5dc"

        # almacena el último resumen (dict) para pre-carga y guardado
        self.ultimo_resumen = {}

        # notebook (pestañas)
        self.nb = ttk.Notebook(master)
        self.nb.pack(fill="both", expand=True, padx=10, pady=8)

        # crear pestañas
        self._build_tab_calculo()
        self._build_tab_costos()
        self._build_tab_guardar()

        # inicializar estado (campos vacios)
        self._set_initial_empty()

    # -----------------------------
    # PESTAÑA CÁLCULO
    # -----------------------------
    def _build_tab_calculo(self):
        frame = ttk.Frame(self.nb)
        self.nb.add(frame, text="Cálculo de Tela")

        # panel beige principal (más amplio)
        panel = tk.Frame(frame, bg=self.panel_bg, bd=1, relief="ridge")
        panel.place(x=12, y=12, width=840, height=540)

        # Título interior
        tk.Label(panel, text="Datos del molde y tela", bg=self.panel_bg, font=("Arial", 11, "bold")).place(x=12, y=8)

        # Radios de modo
        self.modo_var = tk.StringVar(value="cantidad")
        r1 = ttk.Radiobutton(panel, text="Calcular tela según cantidad de objetos/piezas", variable=self.modo_var, value="cantidad", command=self._on_modo_change)
        r1.place(x=12, y=36)
        r2 = ttk.Radiobutton(panel, text="Calcular moldes con X cm de tela", variable=self.modo_var, value="con_tela", command=self._on_modo_change)
        r2.place(x=12, y=60)

        # Campos (lado izquierdo)
        lbl_x = 12
        val_x = 200
        gap_y = 36
        y = 90

        tk.Label(panel, text="Ancho utilizable de tela (cm):", bg=self.panel_bg).place(x=lbl_x, y=y)
        self.entry_ancho_tela = ttk.Entry(panel, width=14)
        self.entry_ancho_tela.place(x=val_x, y=y)

        y += gap_y
        tk.Label(panel, text="Ancho del molde (cm):", bg=self.panel_bg).place(x=lbl_x, y=y)
        self.entry_ancho_molde = ttk.Entry(panel, width=14)
        self.entry_ancho_molde.place(x=val_x, y=y)

        y += gap_y
        tk.Label(panel, text="Alto del molde (cm):", bg=self.panel_bg).place(x=lbl_x, y=y)
        self.entry_alto_molde = ttk.Entry(panel, width=14)
        self.entry_alto_molde.place(x=val_x, y=y)

        y += gap_y
        tk.Label(panel, text="Margen de costura por lado (cm):", bg=self.panel_bg).place(x=lbl_x, y=y)
        self.entry_margen = ttk.Entry(panel, width=14)
        self.entry_margen.place(x=val_x, y=y)

        y += gap_y
        tk.Label(panel, text="Desperdicio (%):", bg=self.panel_bg).place(x=lbl_x, y=y)
        self.entry_desperdicio = ttk.Entry(panel, width=14)
        self.entry_desperdicio.place(x=val_x, y=y)

        # Campos derecho (cantidad / largo disponible)
        ry = 72
        rx = 420
        tk.Label(panel, text="Cantidad de objetos/piezas:", bg=self.panel_bg).place(x=rx, y=ry)
        self.entry_cantidad = ttk.Entry(panel, width=14)
        self.entry_cantidad.place(x=rx + 200, y=ry)

        ry += gap_y
        tk.Label(panel, text="Largo de tela disponible (cm):", bg=self.panel_bg).place(x=rx, y=ry)
        self.entry_largo_tela = ttk.Entry(panel, width=14)
        self.entry_largo_tela.place(x=rx + 200, y=ry)

        # Checkbox doble molde (inicia destildado)
        ry += gap_y
        self.var_doble = tk.BooleanVar(value=False)
        self.chk_doble = ttk.Checkbutton(panel, text="¿El molde se corta por 2? (frente + contrafrente)", variable=self.var_doble)
        self.chk_doble.place(x=rx, y=ry)

        # Botones calcular / limpiar
        btn_calc = ttk.Button(panel, text="Calcular", command=self._accion_calcular)
        btn_calc.place(x=12, y=320, width=120, height=34)
        btn_limpiar = ttk.Button(panel, text="Limpiar", command=self._accion_limpiar)
        btn_limpiar.place(x=150, y=320, width=120, height=34)

        # Área de resultados (más arriba)
        tk.Label(panel, text="Resumen rápido:", bg=self.panel_bg, font=("Arial", 10, "bold")).place(x=12, y=370)
        self.txt_resumen_rapido = tk.Text(panel, height=8, width=96, state="disabled", wrap="word")
        self.txt_resumen_rapido.place(x=12, y=395)

        # set initial mode
        self._on_modo_change()

    def _on_modo_change(self):
        modo = self.modo_var.get()
        if modo == "cantidad":
            # cantidad enabled, largo disabled; checkbox visible
            self.entry_cantidad.configure(state="normal")
            self.entry_largo_tela.configure(state="disabled")
            self.chk_doble.state(['!disabled'])
            # ensure checkbox visible
            self.chk_doble.place_configure()  # ensure it's placed (already)
        else:
            # cantidad disabled, largo enabled; checkbox hidden
            self.entry_cantidad.configure(state="disabled")
            self.entry_largo_tela.configure(state="normal")
            # hide checkbox by moving off-screen or disabling
            self.chk_doble.state(['disabled'])
            # Note: we keep it visually but disabled so user knows it exists (could .place_forget() if wanted)

    def _accion_limpiar(self):
        # vacía todos los inputs y las salidas; deja checkbox destildado
        for e in [self.entry_ancho_tela, self.entry_ancho_molde, self.entry_alto_molde,
                  self.entry_margen, self.entry_desperdicio, self.entry_cantidad, self.entry_largo_tela]:
            e.delete(0, tk.END)
        self.var_doble.set(False)
        self.txt_resumen_rapido.config(state="normal")
        self.txt_resumen_rapido.delete(1.0, tk.END)
        self.txt_resumen_rapido.config(state="disabled")
        # limpiar resumen y costos previos en memoria (no borra pestaña costos)
        self.ultimo_resumen = {}
        self._actualizar_tab_guardar()

    def _accion_calcular(self):
        # leer datos con validación
        try:
            ancho_tela = float(self.entry_ancho_tela.get())
            ancho_molde = float(self.entry_ancho_molde.get())
            alto_molde = float(self.entry_alto_molde.get())
            margen = float(self.entry_margen.get() or '0')
            desperdicio = float(self.entry_desperdicio.get() or '0')
        except ValueError:
            messagebox.showerror("Error", "Completá los campos numéricos correctamente (en cm o %).")
            return

        modo = self.modo_var.get()
        if modo == "cantidad":
            try:
                cantidad = int(self.entry_cantidad.get())
            except ValueError:
                messagebox.showerror("Error", "Ingresá una cantidad válida (entero).")
                return
            doble = bool(self.var_doble.get())
            res, err = calcular_tela_por_cantidad(ancho_tela, ancho_molde, alto_molde,
                                                  margen, desperdicio, cantidad, doble_molde=doble)
            if err:
                messagebox.showerror("Error", err)
                return
            # guardar ultimo resumen
            self.ultimo_resumen = res
            # limpiar costos previos si existen (no obligatorio)
            # mostrar resumen rápido y actualizar pestaña Guardar
            self._mostrar_resumen_rapido(res)
            self._actualizar_tab_guardar()

            # Limpiar y precargar costos
            self.entry_largo_para_costo.delete(0, tk.END)
            self.entry_cantidad_para_costo.delete(0, tk.END)
            self.txt_costos.config(state="normal")
            self.txt_costos.delete(1.0, tk.END)
            self.txt_costos.config(state="disabled")
            self.var_precargar.set(False)

        else:  # modo con_tela
            try:
                largo_disponible = float(self.entry_largo_tela.get())
            except ValueError:
                messagebox.showerror("Error", "Ingresá un largo de tela válido (en cm).")
                return
            res, err = calcular_moldes_con_tela(ancho_tela, ancho_molde, alto_molde,
                                                margen, desperdicio, largo_disponible)
            if err:
                messagebox.showerror("Error", err)
                return
            self.ultimo_resumen = res
            self._mostrar_resumen_rapido(res)
            self._actualizar_tab_guardar()

            # Limpiar y precargar costos
            self.entry_largo_para_costo.delete(0, tk.END)
            self.entry_cantidad_para_costo.delete(0, tk.END)
            self.txt_costos.config(state="normal")
            self.txt_costos.delete(1.0, tk.END)
            self.txt_costos.config(state="disabled")
            self.var_precargar.set(False)

        # al calcular, dejamos la pestaña Guardar con la info, costo no se calcula automáticamente
        # (el usuario puede ir a Costos y precargar si desea)

    def _mostrar_resumen_rapido(self, resdict):
        # construye un texto legible y lo muestra en el cuadro rapido
        lines = []
        modo = resdict.get("modo", "")
        lines.append(f"Modo: {modo}")
        # medidas y totales
        if "ancho_molde_total_cm" in resdict:
            lines.append(f"Ancho molde total: {cm_to_m_str(resdict['ancho_molde_total_cm'])}")
        if "alto_molde_total_cm" in resdict:
            lines.append(f"Alto molde total: {cm_to_m_str(resdict['alto_molde_total_cm'])}")
        if "moldes_por_fila" in resdict:
            lines.append(f"Moldes por fila: {resdict['moldes_por_fila']}")
        if "filas_necesarias" in resdict:
            lines.append(f"Filas necesarias: {resdict['filas_necesarias']}")
        if "cantidad_solicitada" in resdict:
            lines.append(f"Cantidad solicitada (considerando doble si aplica): {resdict['cantidad_solicitada']}")
        if "total_moldes_obtenibles" in resdict:
            lines.append(f"Total moldes obtenibles: {resdict['total_moldes_obtenibles']}")
        if "largo_total_con_desperdicio_cm" in resdict:
            lines.append(f"Largo total necesario (con desperdicio): {cm_to_m_str(resdict['largo_total_con_desperdicio_cm'])}")
        if "largo_utilizable_cm" in resdict:
            lines.append(f"Largo utilizable descontando desperdicio: {cm_to_m_str(resdict['largo_utilizable_cm'])}")
        text = "\n".join(lines)
        self.txt_resumen_rapido.config(state="normal")
        self.txt_resumen_rapido.delete(1.0, tk.END)
        self.txt_resumen_rapido.insert(tk.END, text)
        self.txt_resumen_rapido.config(state="disabled")

    # -----------------------------
    # PESTAÑA COSTOS
    # -----------------------------
    def _build_tab_costos(self):
        frame = ttk.Frame(self.nb)
        self.nb.add(frame, text="Costos")

        panel = tk.Frame(frame, bg=self.panel_bg, bd=1, relief="ridge")
        panel.place(x=12, y=12, width=840, height=220)
        tk.Label(panel, text="Costos (manual o precargados desde último cálculo)", bg=self.panel_bg, font=("Arial", 11, "bold")).place(x=12, y=8)

        # campos (inician vacíos)
        tk.Label(panel, text="Largo de tela para costo (cm):", bg=self.panel_bg).place(x=12, y=44)
        self.entry_largo_para_costo = ttk.Entry(panel, width=16)
        self.entry_largo_para_costo.place(x=220, y=44)

        tk.Label(panel, text="Precio por metro:", bg=self.panel_bg).place(x=12, y=84)
        self.entry_precio_metro = ttk.Entry(panel, width=16)
        self.entry_precio_metro.place(x=220, y=84)

        tk.Label(panel, text="Cantidad de objetos (para costo unitario):", bg=self.panel_bg).place(x=12, y=124)
        self.entry_cantidad_para_costo = ttk.Entry(panel, width=16)
        self.entry_cantidad_para_costo.place(x=320, y=124)

        # Checkbox precargar desde último cálculo (vacío al inicio)
        self.var_precargar = tk.BooleanVar(value=False)
        self.chk_precargar = ttk.Checkbutton(panel, text="Precargar datos del último cálculo (si existen)", variable=self.var_precargar, command=self._on_precargar_toggle)
        self.chk_precargar.place(x=12, y=164)

        # botón calcular costos y área de texto
        btn_calc_cost = ttk.Button(frame, text="Calcular costos", command=self._accion_calcular_costos)
        btn_calc_cost.place(x=12, y=252, width=140, height=34)

        btn_limpiar_costos = ttk.Button(frame, text="Limpiar", command=self._accion_limpiar_costos)
        btn_limpiar_costos.place(x=160, y=252, width=120, height=34)

        self.txt_costos = tk.Text(frame, height=8, width=100, state="disabled", wrap="word")
        self.txt_costos.place(x=12, y=300)

    def _on_precargar_toggle(self):
        if not self.var_precargar.get():
            # si destilda, no hace nada (no borra)
            return
        # si hay ultimo resumen, precarga campos SOLO si estan vacíos
        if not self.ultimo_resumen:
            messagebox.showinfo("Info", "No hay cálculo previo para precargar.")
            self.var_precargar.set(False)
            return
        res = self.ultimo_resumen
        # si hay un valor de largo calculado, precargar en entry_largo_para_costo si vacío
        if "largo_total_con_desperdicio_cm" in res:
            if not self.entry_largo_para_costo.get():
                self.entry_largo_para_costo.insert(0, str(format_number(res["largo_total_con_desperdicio_cm"])))
            if "cantidad_solicitada" in res and not self.entry_cantidad_para_costo.get():
                self.entry_cantidad_para_costo.insert(0, str(res["cantidad_solicitada"]))
        elif "largo_utilizable_cm" in res:
            # en modo 'con_tela', podemos precargar el largo utilizable o el largo disponible original
            if not self.entry_largo_para_costo.get():
                largo_val = res.get("largo_utilizable_cm", res.get("largo_tela_disponible_cm", ""))
                self.entry_largo_para_costo.insert(0, str(format_number(largo_val)))
            if "total_moldes_obtenibles" in res and not self.entry_cantidad_para_costo.get():
                self.entry_cantidad_para_costo.insert(0, str(res["total_moldes_obtenibles"]))

    def _accion_calcular_costos(self):
        try:
            largo_cm = float(self.entry_largo_para_costo.get())
            precio_metro = float(self.entry_precio_metro.get())
            cantidad = int(self.entry_cantidad_para_costo.get())
        except ValueError:
            messagebox.showerror("Error", "Completá los campos de costo correctamente (números).")
            return
        costos = calcular_costos_desde_largo(largo_cm, precio_metro, cantidad_unidades=cantidad)
        # mostrar
        self.txt_costos.config(state="normal")
        self.txt_costos.delete(1.0, tk.END)
        self.txt_costos.insert(tk.END, f"Largo usado (cm): {format_cost(largo_cm)}\n")
        self.txt_costos.insert(tk.END, f"Precio por metro: {format_cost(precio_metro)}\n")
        self.txt_costos.insert(tk.END, f"Costo total: {format_cost(costos['costo_total'])}\n")
        self.txt_costos.insert(tk.END, f"Costo por unidad: {format_cost(costos['costo_unitario'])}\n")
        self.txt_costos.config(state="disabled")
        # añadir costos al ultimo resumen (para mostrar y guardar)
        self.ultimo_resumen.update({
            "precio_por_metro": precio_metro,
            "largo_para_costo_cm": largo_cm,
            "costo_total": costos["costo_total"],
            "costo_unitario": costos["costo_unitario"]
        })
        self._actualizar_tab_guardar()

    def _accion_limpiar_costos(self):
        self.entry_largo_para_costo.delete(0, tk.END)
        self.entry_precio_metro.delete(0, tk.END)
        self.entry_cantidad_para_costo.delete(0, tk.END)
        self.txt_costos.config(state="normal")
        self.txt_costos.delete(1.0, tk.END)
        self.txt_costos.config(state="disabled")

    # -----------------------------
    # PESTAÑA GUARDAR / RESUMEN VISUAL
    # -----------------------------
    def _build_tab_guardar(self):
        frame = ttk.Frame(self.nb)
        self.nb.add(frame, text="Guardar Resultados")

        panel = tk.Frame(frame, bg=self.panel_bg, bd=1, relief="ridge")
        panel.place(x=12, y=12, width=840, height=540)
        tk.Label(panel, text="Resumen (revisá antes de guardar)", bg=self.panel_bg, font=("Arial", 11, "bold")).place(x=12, y=8)

        # Estilo para Treeview
        style = ttk.Style()
        style.configure("Treeview.Heading", font=(None, 10, "bold"))

        # Treeview para mostrar la tabla campo-valor (más visual)
        cols = ("campo", "valor")
        self.tree_resumen = ttk.Treeview(panel, columns=cols, show="headings", height=14)
        self.tree_resumen.heading("campo", text="Campo")
        self.tree_resumen.heading("valor", text="Valor")
        self.tree_resumen.column("campo", width=360, anchor="w")
        self.tree_resumen.column("valor", width=420, anchor="w")
        self.tree_resumen.place(x=12, y=44)
        self.tree_resumen.tag_configure("bold", font=(None, 10, "bold"))

        # Campo de notas
        tk.Label(panel, text="Notas (opcional):", bg=self.panel_bg).place(x=12, y=380)
        self.txt_notas = tk.Text(panel, height=4, width=96, wrap="word")
        self.txt_notas.place(x=12, y=405)

        # Botón guardar (nombre sugerido editable en diálogo)
        btn_guardar = ttk.Button(frame, text="💾 Guardar resultados", command=self._accion_guardar_dialog)
        btn_guardar.place(x=650, y=500, width=180, height=36)

    def _actualizar_tab_guardar(self):
        # limpia y vuelve a poblar el Treeview con el ultimo_resumen en formato amigable
        for r in self.tree_resumen.get_children():
            self.tree_resumen.delete(r)
        if not self.ultimo_resumen:
            return

        # Etiquetas amigables
        friendly_labels = {
            "modo": "Modo de cálculo",
            "ancho_tela_cm": "Ancho de tela (cm)",
            "ancho_molde_cm": "Ancho del molde (cm)",
            "alto_molde_cm": "Alto del molde (cm)",
            "margen_costura_cm_por_lado": "Margen de costura por lado (cm)",
            "ancho_molde_total_cm": "Ancho del molde total (cm)",
            "alto_molde_total_cm": "Alto del molde total (cm)",
            "moldes_por_fila": "Moldes por fila",
            "filas_necesarias": "Filas necesarias",
            "cantidad_solicitada": "Cantidad solicitada",
            "doble_molde": "Molde doble (frente y contrafrente)",
            "largo_total_sin_desperdicio_cm": "Largo total sin desperdicio (cm)",
            "largo_total_con_desperdicio_cm": "Largo total con desperdicio (cm)",
            "largo_tela_disponible_cm": "Largo de tela disponible (cm)",
            "largo_utilizable_cm": "Largo utilizable (cm)",
            "filas_posibles": "Filas posibles",
            "total_moldes_obtenibles": "Total de moldes obtenibles",
            "precio_por_metro": "Precio por metro",
            "largo_para_costo_cm": "Largo para costo (cm)",
            "costo_total": "Costo total",
            "costo_unitario": "Costo unitario"
        }

        # transformar campos a texto legible y agregar conversión a metros donde aplique
        def add_row(k, v, is_bold=False):
            label = friendly_labels.get(k, k)
            display_val = v

            if k == "doble_molde":
                display_val = "Requiere" if v else "No requiere"
            # si es medida en cm, añadimos la versión en m entre paréntesis
            elif isinstance(v, (int, float)) and ("cm" in (str(k).lower()) or "largo" in str(k).lower() or "ancho" in str(k).lower() or "alto" in str(k).lower()):
                display_val = f"{format_number(v)} cm ({float(v)/100:.2f} m)"
            else:
                display_val = str(v)

            tags = ("bold",) if is_bold else ()
            self.tree_resumen.insert("", tk.END, values=(label, display_val), tags=tags)

        # orden preferido de campos para mostrar (inteligente)
        pref = [
            "modo", "ancho_tela_cm", "ancho_molde_cm", "alto_molde_cm",
            "margen_costura_cm_por_lado", "ancho_molde_total_cm", "alto_molde_total_cm",
            "moldes_por_fila", "filas_necesarias", "cantidad_solicitada",
            "largo_total_sin_desperdicio_cm", "largo_total_con_desperdicio_cm",
            "largo_tela_disponible_cm", "largo_utilizable_cm",
            "filas_posibles", "total_moldes_obtenibles",
            "precio_por_metro", "largo_para_costo_cm", "costo_total", "costo_unitario"
        ]
        added = set()
        for i, key in enumerate(pref):
            if key in self.ultimo_resumen:
                add_row(key, self.ultimo_resumen[key], is_bold=(i == 0))
                added.add(key)
        # añadir el resto de campos no listados
        for k, v in self.ultimo_resumen.items():
            if k not in added:
                add_row(k, v)

    def _accion_guardar_dialog(self):
        if not self.ultimo_resumen:
            messagebox.showinfo("Info", "No hay resultados para guardar. Primero realizá un cálculo.")
            return
        # sugerir nombre con fecha/hora, pero dejar editable en diálogo
        ahora = datetime.now().strftime("%Y-%m-%d_%H%M")
        sugerido = f"resultados_tela_{ahora}"
        filepath = filedialog.asksaveasfilename(title="Guardar resultados como",
                                                initialfile=sugerido,
                                                defaultextension=".xlsx",
                                                filetypes=[("Excel (.xlsx)", "*.xlsx"), ("Texto (.txt)", "*.txt")])
        if not filepath:
            return
        try:
            # preparar un resumen legible para guardar (campo -> valor, con cm->m formateo)
            resumen_guardado = {}
            # Usar las mismas etiquetas amigables que en la tabla
            for child_id in self.tree_resumen.get_children():
                item = self.tree_resumen.item(child_id)
                campo, valor = item['values']
                resumen_guardado[campo] = valor

            # Añadir notas
            notas = self.txt_notas.get("1.0", tk.END).strip()
            if notas:
                resumen_guardado["Notas"] = notas

            if filepath.lower().endswith(".txt"):
                guardar_txt(filepath, resumen_guardado)
            else:
                guardar_xlsx(filepath, resumen_guardado)
            messagebox.showinfo("Guardado", f"Archivo guardado en:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Error al guardar", str(e))


    # -----------------------------
    # INICIALIZACIONES
    # -----------------------------
    def _set_initial_empty(self):
        # Dejar todos los campos vacíos al inicio
        for e in [self.entry_ancho_tela, self.entry_ancho_molde, self.entry_alto_molde,
                  self.entry_margen, self.entry_desperdicio, self.entry_cantidad, self.entry_largo_tela,
                  self.entry_largo_para_costo, self.entry_precio_metro, self.entry_cantidad_para_costo]:
            try:
                e.delete(0, tk.END)
            except Exception:
                pass
        # checkbox desmarcado
        self.var_doble.set(False)
        self.var_precargar = getattr(self, "var_precargar", tk.BooleanVar(value=False))
        # limpiar texto y tabla
        self.txt_resumen_rapido.config(state="normal")
        self.txt_resumen_rapido.delete(1.0, tk.END)
        self.txt_resumen_rapido.config(state="disabled")
        if hasattr(self, "txt_costos"):
            self.txt_costos.config(state="normal")
            self.txt_costos.delete(1.0, tk.END)
            self.txt_costos.config(state="disabled")
        if hasattr(self, "tree_resumen"):
            for r in self.tree_resumen.get_children():
                self.tree_resumen.delete(r)


# ------------------------
# RUN
# ------------------------
if __name__ == "__main__":
    root = tk.Tk()
    app = CalculadoraTelaApp(root)
    root.mainloop()
